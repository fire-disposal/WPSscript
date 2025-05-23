#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
将Excel表格中指定列的公式单元格改为它们自身计算得出的结果（固定值）
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 文件读取部分，便于修改需读取文件名
input_file = "example.xlsx"  # 请修改为实际的文件名

# 需要处理的列配置
# 可以使用列字母（如 'A', 'B', 'C'）或列范围（如 'A:C'）
columns_to_process = [
    "B",       # 单列
    "D:F",     # 列范围
]

def parse_column_reference(column_ref):
    """
    解析列引用，支持单列和列范围
    
    Args:
        column_ref: 列引用，如 'A' 或 'A:C'
    
    Returns:
        列索引列表
    """
    if ":" in column_ref:
        start_col, end_col = column_ref.split(":")
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        return list(range(start_idx, end_idx + 1))
    else:
        return [column_index_from_string(column_ref)]

def convert_formulas_to_values(excel_path, columns):
    """
    将指定列的公式单元格转换为它们的计算结果（固定值）
    
    Args:
        excel_path: Excel文件路径
        columns: 要处理的列配置列表
    
    Returns:
        处理后的工作簿对象
    """
    print(f"正在处理文件: {excel_path}")
    
    # 检查文件是否存在
    if not os.path.exists(excel_path):
        print(f"错误: 文件 '{excel_path}' 不存在!")
        return None
    
    try:
        # 打开工作簿，保留公式以便我们可以获取它们的值
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        
        # 同时打开一个data_only=True的工作簿，用于获取公式计算结果
        wb_values = openpyxl.load_workbook(excel_path, data_only=True)
        
        # 解析所有要处理的列索引
        column_indices = []
        for col_ref in columns:
            column_indices.extend(parse_column_reference(col_ref))
        
        # 去重并排序
        column_indices = sorted(set(column_indices))
        
        # 处理每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_values = wb_values[sheet_name]
            
            print(f"正在处理工作表: {sheet_name}")
            
            # 跟踪已处理的公式单元格数量
            formula_count = 0
            
            # 遍历每一行
            for row in range(1, ws.max_row + 1):
                # 只处理指定的列
                for col_idx in column_indices:
                    cell = ws.cell(row=row, column=col_idx)
                    
                    # 检查单元格是否包含公式
                    if cell.data_type == 'f':
                        # 获取对应的计算结果
                        value_cell = ws_values.cell(row=row, column=col_idx)
                        
                        # 保存原始公式以便输出
                        original_formula = cell.value
                        
                        # 将公式替换为计算结果
                        cell.value = value_cell.value
                        
                        # 保持原始格式
                        cell.number_format = value_cell.number_format
                        
                        formula_count += 1
                        
                        # 输出详细信息（每10个公式输出一次，避免输出过多）
                        if formula_count % 10 == 0 or formula_count < 10:
                            col_letter = get_column_letter(col_idx)
                            print(f"  - 已处理: {sheet_name}!{col_letter}{row} - 公式 '{original_formula}' 替换为值 '{cell.value}'")
            
            print(f"  - 工作表 '{sheet_name}' 中共处理了 {formula_count} 个公式单元格")
        
        return wb
    
    except Exception as e:
        print(f"处理过程中出错: {str(e)}")
        return None

def main():
    # 构建输出文件名
    file_name, file_ext = os.path.splitext(input_file)
    output_file = f"{file_name}（已修改）{file_ext}"
    
    # 执行公式转换
    wb = convert_formulas_to_values(input_file, columns_to_process)
    
    if wb:
        # 保存修改后的工作簿
        wb.save(output_file)
        print(f"处理完成! 工作簿已保存为: {output_file}")
        print(f"注意: 所有指定列中的公式已被替换为它们的计算结果（固定值）")

if __name__ == "__main__":
    main()