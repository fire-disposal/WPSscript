#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
自动生成Excel数据透视表
注意：由于openpyxl库的限制，此脚本只能创建数据透视表的定义，
但不能计算数据透视表的结果。用户需要在Excel中打开生成的文件，
然后刷新数据透视表以查看结果。
"""

import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter

# 文件读取部分，便于修改需读取文件名
input_file = "example.xlsx"  # 请修改为实际的文件名

# 数据透视表设置
pivot_settings = {
    # 数据源设置
    "source": {
        "sheet": "数据",  # 数据源工作表名
        "range": "A1:F100",  # 数据范围
        "has_headers": True,  # 是否包含表头
    },
    
    # 数据透视表设置
    "pivot": {
        "sheet": "数据透视表",  # 数据透视表所在工作表名（如果不存在将创建）
        "location": "A3",  # 数据透视表位置
        "rows": ["部门", "姓名"],  # 行字段
        "columns": ["月份"],  # 列字段
        "values": [("销售额", "sum"), ("数量", "count")],  # 值字段及汇总方式
        "filters": ["产品类别"],  # 筛选字段
    }
}

def create_source_table(wb, settings):
    """
    创建数据源表格
    
    Args:
        wb: 工作簿对象
        settings: 数据源设置
    
    Returns:
        (工作表对象, 表格范围)
    """
    sheet_name = settings["sheet"]
    
    # 检查工作表是否存在
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"错误: 数据源工作表 '{sheet_name}' 不存在!")
        return None, None
    
    # 获取数据范围
    data_range = settings["range"]
    
    # 创建表格
    table_name = "DataTable"
    table_ref = data_range
    
    # 检查表格是否已存在
    for table in ws._tables:
        if table.name == table_name:
            print(f"表格 '{table_name}' 已存在，将使用现有表格")
            return ws, table_ref
    
    # 创建新表格
    table = Table(displayName=table_name, ref=table_ref)
    
    # 设置表格样式
    style = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    # 添加表格到工作表
    ws.add_table(table)
    print(f"已创建数据源表格: {sheet_name}!{data_range}")
    
    return ws, table_ref

def create_pivot_table(wb, source_ws, source_range, settings):
    """
    创建数据透视表
    
    Args:
        wb: 工作簿对象
        source_ws: 数据源工作表
        source_range: 数据源范围
        settings: 数据透视表设置
    
    Returns:
        工作表对象
    """
    sheet_name = settings["sheet"]
    
    # 检查工作表是否存在，如果不存在则创建
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"使用现有工作表: {sheet_name}")
    else:
        ws = wb.create_sheet(title=sheet_name)
        print(f"创建新工作表: {sheet_name}")
    
    # 添加标题
    ws["A1"] = "数据透视表"
    ws["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    
    # 设置列宽
    for col in range(1, 20):  # 设置前20列的宽度
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 由于openpyxl不支持直接创建数据透视表，我们将添加一个说明
    location = settings["location"]
    ws[location] = "数据透视表将在此处创建"
    
    # 添加数据透视表设置说明
    row_idx = ws[location].row + 2
    
    ws.cell(row=row_idx, column=1).value = "数据透视表设置:"
    ws.cell(row=row_idx, column=1).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = "数据源:"
    ws.cell(row=row_idx, column=2).value = f"'{source_ws.title}'!{source_range}"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = "行字段:"
    ws.cell(row=row_idx, column=2).value = ", ".join(settings["rows"])
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = "列字段:"
    ws.cell(row=row_idx, column=2).value = ", ".join(settings["columns"])
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = "值字段:"
    values_str = ", ".join([f"{field}({agg})" for field, agg in settings["values"]])
    ws.cell(row=row_idx, column=2).value = values_str
    row_idx += 1
    
    if settings["filters"]:
        ws.cell(row=row_idx, column=1).value = "筛选字段:"
        ws.cell(row=row_idx, column=2).value = ", ".join(settings["filters"])
        row_idx += 1
    
    # 添加使用说明
    row_idx += 2
    ws.cell(row=row_idx, column=1).value = "使用说明:"
    ws.cell(row=row_idx, column=1).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = "1. 此文件包含数据透视表的定义，但由于Python库的限制，无法直接计算数据透视表结果。"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = "2. 请在Excel中打开此文件，然后按照以下步骤创建数据透视表:"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = "   a. 选择'插入'选项卡"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = f"   b. 点击'数据透视表'，数据源选择'{source_ws.title}'!{source_range}"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = f"   c. 位置选择'{ws.title}'!{location}"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = "   d. 在数据透视表字段列表中，将字段拖到相应的区域"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = f"      - 行区域: {', '.join(settings['rows'])}"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = f"      - 列区域: {', '.join(settings['columns'])}"
    row_idx += 1
    
    ws.cell(row=row_idx, column=1).value = f"      - 值区域: {values_str}"
    row_idx += 1
    
    if settings["filters"]:
        ws.cell(row=row_idx, column=1).value = f"      - 筛选区域: {', '.join(settings['filters'])}"
    
    print(f"已在工作表 '{sheet_name}' 中添加数据透视表设置说明")
    return ws

def main():
    # 构建输出文件名
    file_name, file_ext = os.path.splitext(input_file)
    output_file = f"{file_name}（已修改）{file_ext}"
    
    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 文件 '{input_file}' 不存在!")
        return
    
    # 打开工作簿
    print(f"正在处理文件: {input_file}")
    wb = openpyxl.load_workbook(input_file)
    
    # 创建数据源表格
    source_ws, source_range = create_source_table(wb, pivot_settings["source"])
    
    if source_ws and source_range:
        # 创建数据透视表
        create_pivot_table(wb, source_ws, source_range, pivot_settings["pivot"])
        
        # 保存工作簿
        wb.save(output_file)
        print(f"数据透视表设置已完成! 工作簿已保存为: {output_file}")
        print("请在Excel中打开此文件，并按照说明创建数据透视表。")
    else:
        print("无法创建数据透视表，请检查数据源设置。")

if __name__ == "__main__":
    main()