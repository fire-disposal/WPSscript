#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
快速填充Excel表格中的空值
- 可自定义选择工作表
- 可自定义选择区域（单元格、行、列或区域范围）
- 可自定义空值匹配条件（None、空字符串、空格字符、数字0等）
- 可自定义填充值
"""

import os
import re
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

# 文件读取部分，便于修改需读取文件名
input_file = "example.xlsx"  # 请修改为实际的文件名

# 需要处理的工作表配置
# 可以指定工作表名称列表，或使用 None 表示处理所有工作表
worksheets_to_process = [
    "Sheet1",   # 处理名为"Sheet1"的工作表
    "数据表",    # 处理名为"数据表"的工作表
    # None,     # 取消注释此行将处理所有工作表
]

# 需要处理的区域配置
# 可以使用单元格引用（如 'A1'）、行范围（如 '1:5'）、列范围（如 'A:C'）或区域范围（如 'A1:C5'）
areas_to_process = [
    "B:B",      # B列
    "D1:F10",   # D1到F10区域
    "5:10",     # 第5行到第10行
]

# 空值匹配条件配置
# 可以自定义多种空值匹配条件
empty_value_conditions = {
    "none": True,           # 匹配None值
    "empty_string": True,   # 匹配空字符串 ""
    "whitespace": True,     # 匹配只包含空格的字符串，如 " ", "  " 等
    "zero": False,          # 匹配数字0
    "zero_string": False,   # 匹配字符串 "0"
    "custom_values": [],    # 自定义匹配值列表，如 ["N/A", "#N/A", "NULL"]
    "custom_pattern": None, # 自定义正则表达式模式，如 r"^(NA|N/A)$"
}

# 填充值配置
fill_value = 0  # 用于填充空值的值

def parse_area_reference(area_ref):
    """
    解析区域引用，支持单元格、行范围、列范围和区域范围
    
    Args:
        area_ref: 区域引用，如 'A1'、'1:5'、'A:C' 或 'A1:C5'
    
    Returns:
        (min_row, min_col, max_row, max_col) 元组
    """
    # 检查是否是列范围（如 'A:C'）
    if re.match(r'^[A-Z]+:[A-Z]+$', area_ref):
        start_col, end_col = area_ref.split(":")
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        return (1, start_idx, 1000000, end_idx)  # 使用一个足够大的行数
    
    # 检查是否是行范围（如 '1:5'）
    elif re.match(r'^\d+:\d+$', area_ref):
        start_row, end_row = map(int, area_ref.split(":"))
        return (start_row, 1, end_row, 16384)  # 使用Excel的最大列数
    
    # 检查是否是区域范围（如 'A1:C5'）或单元格（如 'A1'）
    else:
        try:
            # 尝试使用openpyxl的range_boundaries函数解析
            return range_boundaries(area_ref)
        except:
            # 如果解析失败，可能是无效的引用
            print(f"警告: 无法解析区域引用 '{area_ref}'，将被跳过")
            return None

def is_empty_value(value, conditions):
    """
    根据配置的条件检查值是否为"空值"
    
    Args:
        value: 要检查的值
        conditions: 空值匹配条件配置
    
    Returns:
        如果值匹配任何空值条件，则返回True，否则返回False
    """
    # 检查None值
    if conditions.get("none") and value is None:
        return True
    
    # 对于非None值，转换为字符串进行进一步检查
    if value is not None:
        # 检查空字符串
        if conditions.get("empty_string") and value == "":
            return True
        
        # 检查只包含空格的字符串
        if conditions.get("whitespace") and isinstance(value, str) and value.strip() == "":
            return True
        
        # 检查数字0
        if conditions.get("zero") and value == 0:
            return True
        
        # 检查字符串"0"
        if conditions.get("zero_string") and value == "0":
            return True
        
        # 检查自定义值列表
        custom_values = conditions.get("custom_values", [])
        if custom_values and value in custom_values:
            return True
        
        # 检查自定义正则表达式模式
        custom_pattern = conditions.get("custom_pattern")
        if custom_pattern and isinstance(value, str) and re.match(custom_pattern, value):
            return True
    
    return False

def fill_empty_cells(excel_path, worksheets, areas, empty_conditions, fill_value):
    """
    填充指定工作表和区域中的空值单元格
    
    Args:
        excel_path: Excel文件路径
        worksheets: 要处理的工作表名称列表，None表示处理所有工作表
        areas: 要处理的区域配置列表
        empty_conditions: 空值匹配条件配置
        fill_value: 用于填充的值
    
    Returns:
        处理后的工作簿对象
    """
    print(f"正在处理文件: {excel_path}")
    
    # 检查文件是否存在
    if not os.path.exists(excel_path):
        print(f"错误: 文件 '{excel_path}' 不存在!")
        return None
    
    try:
        # 打开工作簿
        wb = openpyxl.load_workbook(excel_path)
        
        # 解析所有要处理的区域
        parsed_areas = []
        for area_ref in areas:
            area_bounds = parse_area_reference(area_ref)
            if area_bounds:
                parsed_areas.append((area_ref, area_bounds))
        
        # 确定要处理的工作表
        if worksheets is None or None in worksheets:
            # 如果worksheets为None或包含None，处理所有工作表
            sheets_to_process = wb.sheetnames
            print("将处理所有工作表")
        else:
            # 否则只处理指定的工作表
            sheets_to_process = []
            for sheet_name in worksheets:
                if sheet_name in wb.sheetnames:
                    sheets_to_process.append(sheet_name)
                else:
                    print(f"警告: 工作表 '{sheet_name}' 不存在，将被跳过")
            
            if not sheets_to_process:
                print("错误: 没有找到任何指定的工作表!")
                return wb
        
        # 处理每个工作表
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            
            print(f"正在处理工作表: {sheet_name}")
            
            # 跟踪已填充的单元格数量
            filled_count = 0
            
            # 处理每个区域
            for area_ref, (min_row, min_col, max_row, max_col) in parsed_areas:
                print(f"  - 处理区域: {area_ref}")
                
                # 确保不超过工作表的实际大小
                max_row = min(max_row, ws.max_row)
                max_col = min(max_col, ws.max_column)
                
                # 遍历区域中的每个单元格
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        
                        # 检查单元格值是否为"空值"
                        if is_empty_value(cell.value, empty_conditions):
                            # 记录原始值以便输出
                            original_value = cell.value
                            
                            # 填充空值
                            cell.value = fill_value
                            
                            filled_count += 1
                            
                            # 输出详细信息（每10个单元格输出一次，避免输出过多）
                            if filled_count % 10 == 0 or filled_count < 10:
                                col_letter = get_column_letter(col)
                                print(f"    - 已填充: {sheet_name}!{col_letter}{row} - 原值 '{original_value}' 替换为 '{fill_value}'")
                
            print(f"  - 工作表 '{sheet_name}' 中共填充了 {filled_count} 个空值单元格")
        
        return wb
    
    except Exception as e:
        print(f"处理过程中出错: {str(e)}")
        return None

def main():
    # 构建输出文件名
    file_name, file_ext = os.path.splitext(input_file)
    output_file = f"{file_name}（已修改）{file_ext}"
    
    # 执行空值填充
    wb = fill_empty_cells(input_file, worksheets_to_process, areas_to_process, empty_value_conditions, fill_value)
    
    if wb:
        # 保存修改后的工作簿
        wb.save(output_file)
        print(f"处理完成! 工作簿已保存为: {output_file}")
        print(f"注意: 所有指定区域中的空值已被填充为: {fill_value}")

if __name__ == "__main__":
    main()