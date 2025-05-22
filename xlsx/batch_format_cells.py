#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
批量格式化Excel单元格
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_PERCENTAGE_00

# 文件读取部分，便于修改需读取文件名
input_file = "example.xlsx"  # 请修改为实际的文件名

# 格式化设置
format_settings = {
    # 格式化区域，格式为 "工作表名:起始单元格:结束单元格"
    "区域1": {
        "range": "Sheet1:A1:G10",
        "header_row": True,  # 是否包含表头行
        "number_format": "#,##0.00",  # 数字格式
        "alignment": {"horizontal": "center", "vertical": "center"},  # 对齐方式
        "font": {"name": "Arial", "size": 11, "bold": False},  # 字体设置
        "header_font": {"name": "Arial", "size": 12, "bold": True},  # 表头字体设置
        "header_fill": {"type": "solid", "color": "DDEBF7"},  # 表头填充颜色
        "border": {"style": "thin", "color": "000000"},  # 边框设置
        "auto_filter": True,  # 是否添加筛选
        "freeze_panes": "A2",  # 冻结窗格位置
    },
    # 可以添加更多区域的格式化设置
    "区域2": {
        "range": "Sheet2:B2:E20",
        "header_row": False,
        "number_format": "0.00%",
        "alignment": {"horizontal": "right", "vertical": "center"},
        "font": {"name": "Calibri", "size": 10, "bold": False},
        "border": {"style": "medium", "color": "4472C4"},
        "auto_filter": False,
        "freeze_panes": None,
    },
}

def parse_range(range_str):
    """
    解析单元格范围字符串
    
    Args:
        range_str: 格式为 "工作表名:起始单元格:结束单元格"
    
    Returns:
        (工作表名, 起始单元格, 结束单元格)
    """
    parts = range_str.split(":")
    if len(parts) != 3:
        raise ValueError(f"无效的范围格式: {range_str}，应为 '工作表名:起始单元格:结束单元格'")
    
    sheet_name = parts[0]
    start_cell = parts[1]
    end_cell = parts[2]
    
    return sheet_name, start_cell, end_cell

def batch_format_cells(excel_path, settings):
    """
    批量格式化Excel单元格
    
    Args:
        excel_path: Excel文件路径
        settings: 格式化设置字典
    
    Returns:
        格式化后的Workbook对象
    """
    print(f"正在处理文件: {excel_path}")
    
    # 检查文件是否存在
    if not os.path.exists(excel_path):
        print(f"错误: 文件 '{excel_path}' 不存在!")
        return None
    
    # 打开工作簿
    wb = openpyxl.load_workbook(excel_path)
    
    # 遍历所有格式化设置
    for area_name, area_settings in settings.items():
        print(f"正在格式化区域: {area_name}")
        
        # 解析单元格范围
        try:
            sheet_name, start_cell, end_cell = parse_range(area_settings["range"])
        except ValueError as e:
            print(f"错误: {str(e)}")
            continue
        
        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            print(f"错误: 工作表 '{sheet_name}' 不存在!")
            continue
        
        # 获取工作表
        ws = wb[sheet_name]
        
        # 获取单元格范围
        cell_range = ws[f"{start_cell}:{end_cell}"]
        
        # 创建边框样式
        if "border" in area_settings:
            border_style = area_settings["border"]["style"]
            border_color = area_settings["border"]["color"]
            border = Border(
                left=Side(style=border_style, color=border_color),
                right=Side(style=border_style, color=border_color),
                top=Side(style=border_style, color=border_color),
                bottom=Side(style=border_style, color=border_color)
            )
        else:
            border = None
        
        # 创建表头字体和填充
        if area_settings.get("header_row", False):
            if "header_font" in area_settings:
                header_font = Font(
                    name=area_settings["header_font"].get("name", "Arial"),
                    size=area_settings["header_font"].get("size", 11),
                    bold=area_settings["header_font"].get("bold", True),
                    italic=area_settings["header_font"].get("italic", False),
                    color=area_settings["header_font"].get("color", "000000")
                )
            else:
                header_font = None
            
            if "header_fill" in area_settings:
                header_fill = PatternFill(
                    fill_type=area_settings["header_fill"].get("type", "solid"),
                    fgColor=area_settings["header_fill"].get("color", "DDEBF7")
                )
            else:
                header_fill = None
        
        # 创建正文字体
        if "font" in area_settings:
            body_font = Font(
                name=area_settings["font"].get("name", "Arial"),
                size=area_settings["font"].get("size", 11),
                bold=area_settings["font"].get("bold", False),
                italic=area_settings["font"].get("italic", False),
                color=area_settings["font"].get("color", "000000")
            )
        else:
            body_font = None
        
        # 创建对齐方式
        if "alignment" in area_settings:
            alignment = Alignment(
                horizontal=area_settings["alignment"].get("horizontal", "general"),
                vertical=area_settings["alignment"].get("vertical", "bottom"),
                wrap_text=area_settings["alignment"].get("wrap_text", False)
            )
        else:
            alignment = None
        
        # 应用格式
        first_row = True
        for row in cell_range:
            for cell in row:
                # 应用边框
                if border:
                    cell.border = border
                
                # 应用对齐方式
                if alignment:
                    cell.alignment = alignment
                
                # 应用字体和填充
                if first_row and area_settings.get("header_row", False):
                    if header_font:
                        cell.font = header_font
                    if header_fill:
                        cell.fill = header_fill
                else:
                    if body_font:
                        cell.font = body_font
                
                # 应用数字格式
                if "number_format" in area_settings and isinstance(cell.value, (int, float)):
                    cell.number_format = area_settings["number_format"]
            
            first_row = False
        
        # 添加筛选
        if area_settings.get("auto_filter", False):
            ws.auto_filter.ref = f"{start_cell}:{end_cell}"
            print(f"  - 已添加筛选: {start_cell}:{end_cell}")
        
        # 冻结窗格
        if area_settings.get("freeze_panes"):
            freeze_cell = area_settings["freeze_panes"]
            ws.freeze_panes = ws[freeze_cell]
            print(f"  - 已冻结窗格: {freeze_cell}")
        
        print(f"  - 已格式化区域: {sheet_name}!{start_cell}:{end_cell}")
    
    return wb

def main():
    # 构建输出文件名
    file_name, file_ext = os.path.splitext(input_file)
    output_file = f"{file_name}（已修改）{file_ext}"
    
    # 执行批量格式化
    wb = batch_format_cells(input_file, format_settings)
    
    if wb:
        # 保存修改后的工作簿
        wb.save(output_file)
        print(f"格式化完成! 工作簿已保存为: {output_file}")

if __name__ == "__main__":
    main()