#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
重新排列Excel工作表的列顺序
按照第一个工作表的列顺序重新排列其他工作表的列
如果后续工作表不存在第一个表中的某列，则插入新列名，列内容为空
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 文件读取部分，便于修改需读取文件名
input_file = "example.xlsx"  # 请修改为实际的文件名

# 工作表设置
sheet_settings = {
    "sheets_to_process": ["Sheet1", "Sheet2", "Sheet3"],  # 要处理的工作表列表，第一个工作表将作为列顺序的参考
    "has_headers": True,  # 是否包含表头行
}

def get_column_order(ws):
    """
    获取工作表的列顺序
    
    Args:
        ws: 工作表对象
    
    Returns:
        列名列表
    """
    # 获取表头行
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    
    # 提取列名
    column_names = []
    for cell in header_row:
        if cell.value is not None:
            column_names.append(cell.value)
        else:
            # 如果列名为空，使用列字母作为名称
            column_names.append(get_column_letter(cell.column))
    
    return column_names

def get_column_data(ws, col_idx):
    """
    获取工作表中指定列的数据
    
    Args:
        ws: 工作表对象
        col_idx: 列索引
    
    Returns:
        列数据列表（不包含表头）
    """
    column_data = []
    for row in ws.iter_rows(min_row=2):  # 从第二行开始（跳过表头）
        if col_idx <= len(row):
            column_data.append(row[col_idx - 1].value)
        else:
            column_data.append(None)
    
    return column_data

def reorder_columns(wb, sheet_names, has_headers=True):
    """
    重新排列工作表的列顺序
    
    Args:
        wb: 工作簿对象
        sheet_names: 要处理的工作表名称列表
        has_headers: 是否包含表头行
    
    Returns:
        修改后的工作簿对象
    """
    if not sheet_names or len(sheet_names) < 2:
        print("错误: 至少需要两个工作表才能进行列重排序!")
        return wb
    
    # 检查所有工作表是否存在
    missing_sheets = [name for name in sheet_names if name not in wb.sheetnames]
    if missing_sheets:
        print(f"错误: 以下工作表不存在: {', '.join(missing_sheets)}")
        return wb
    
    # 获取第一个工作表作为参考
    reference_sheet_name = sheet_names[0]
    reference_ws = wb[reference_sheet_name]
    
    # 获取参考工作表的列顺序
    if not has_headers:
        print("错误: 必须包含表头行才能进行列重排序!")
        return wb
    
    reference_columns = get_column_order(reference_ws)
    print(f"参考工作表 '{reference_sheet_name}' 的列顺序: {reference_columns}")
    
    # 处理其他工作表
    for sheet_name in sheet_names[1:]:
        print(f"\n正在处理工作表: {sheet_name}")
        ws = wb[sheet_name]
        
        # 获取当前工作表的列顺序
        current_columns = get_column_order(ws)
        print(f"当前列顺序: {current_columns}")
        
        # 创建一个新的工作表来存储重排序后的数据
        new_sheet_name = f"{sheet_name}_reordered"
        if new_sheet_name in wb.sheetnames:
            # 如果已存在同名工作表，则删除
            wb.remove(wb[new_sheet_name])
        
        new_ws = wb.create_sheet(title=new_sheet_name)
        
        # 写入表头
        for col_idx, col_name in enumerate(reference_columns, 1):
            new_ws.cell(row=1, column=col_idx, value=col_name)
        
        # 获取数据行数
        data_rows = ws.max_row - 1 if has_headers else ws.max_row
        
        # 处理每一列
        for col_idx, ref_col_name in enumerate(reference_columns, 1):
            # 检查当前工作表是否包含此列
            if ref_col_name in current_columns:
                # 获取当前工作表中此列的索引
                curr_col_idx = current_columns.index(ref_col_name) + 1
                
                # 复制数据
                for row_idx in range(1, data_rows + 1):
                    cell_value = ws.cell(row=row_idx + 1 if has_headers else row_idx, 
                                         column=curr_col_idx).value
                    new_ws.cell(row=row_idx + 1, column=col_idx, value=cell_value)
                
                print(f"  - 已复制列: {ref_col_name}")
            else:
                # 如果当前工作表不包含此列，则插入空列
                print(f"  - 已插入新列: {ref_col_name} (空)")
        
        # 检查当前工作表中是否有参考工作表中不存在的列
        extra_columns = [col for col in current_columns if col not in reference_columns]
        if extra_columns:
            # 在新工作表中添加这些额外的列
            start_col_idx = len(reference_columns) + 1
            for idx, col_name in enumerate(extra_columns, start_col_idx):
                # 写入列名
                new_ws.cell(row=1, column=idx, value=col_name)
                
                # 获取当前工作表中此列的索引
                curr_col_idx = current_columns.index(col_name) + 1
                
                # 复制数据
                for row_idx in range(1, data_rows + 1):
                    cell_value = ws.cell(row=row_idx + 1 if has_headers else row_idx, 
                                         column=curr_col_idx).value
                    new_ws.cell(row=row_idx + 1, column=idx, value=cell_value)
                
                print(f"  - 已添加额外列: {col_name}")
        
        print(f"工作表 '{sheet_name}' 的列已重排序，结果保存在 '{new_sheet_name}'")
    
    return wb

def reorder_columns_in_place(wb, sheet_names, has_headers=True):
    """
    直接在原工作表中重新排列列顺序
    
    Args:
        wb: 工作簿对象
        sheet_names: 要处理的工作表名称列表
        has_headers: 是否包含表头行
    
    Returns:
        修改后的工作簿对象
    """
    if not sheet_names or len(sheet_names) < 2:
        print("错误: 至少需要两个工作表才能进行列重排序!")
        return wb
    
    # 检查所有工作表是否存在
    missing_sheets = [name for name in sheet_names if name not in wb.sheetnames]
    if missing_sheets:
        print(f"错误: 以下工作表不存在: {', '.join(missing_sheets)}")
        return wb
    
    # 获取第一个工作表作为参考
    reference_sheet_name = sheet_names[0]
    reference_ws = wb[reference_sheet_name]
    
    # 获取参考工作表的列顺序
    if not has_headers:
        print("错误: 必须包含表头行才能进行列重排序!")
        return wb
    
    reference_columns = get_column_order(reference_ws)
    print(f"参考工作表 '{reference_sheet_name}' 的列顺序: {reference_columns}")
    
    # 处理其他工作表
    for sheet_name in sheet_names[1:]:
        print(f"\n正在处理工作表: {sheet_name}")
        ws = wb[sheet_name]
        
        # 获取当前工作表的列顺序
        current_columns = get_column_order(ws)
        print(f"当前列顺序: {current_columns}")
        
        # 创建一个临时工作表来存储重排序后的数据
        temp_sheet_name = f"Temp_{sheet_name}"
        if temp_sheet_name in wb.sheetnames:
            # 如果已存在同名工作表，则删除
            wb.remove(wb[temp_sheet_name])
        
        temp_ws = wb.create_sheet(title=temp_sheet_name)
        
        # 写入表头
        for col_idx, col_name in enumerate(reference_columns, 1):
            temp_ws.cell(row=1, column=col_idx, value=col_name)
        
        # 获取数据行数
        data_rows = ws.max_row - 1 if has_headers else ws.max_row
        
        # 处理每一列
        for col_idx, ref_col_name in enumerate(reference_columns, 1):
            # 检查当前工作表是否包含此列
            if ref_col_name in current_columns:
                # 获取当前工作表中此列的索引
                curr_col_idx = current_columns.index(ref_col_name) + 1
                
                # 复制数据
                for row_idx in range(1, data_rows + 1):
                    cell_value = ws.cell(row=row_idx + 1 if has_headers else row_idx, 
                                         column=curr_col_idx).value
                    temp_ws.cell(row=row_idx + 1, column=col_idx, value=cell_value)
                
                print(f"  - 已复制列: {ref_col_name}")
            else:
                # 如果当前工作表不包含此列，则插入空列
                print(f"  - 已插入新列: {ref_col_name} (空)")
        
        # 检查当前工作表中是否有参考工作表中不存在的列
        extra_columns = [col for col in current_columns if col not in reference_columns]
        if extra_columns:
            # 在临时工作表中添加这些额外的列
            start_col_idx = len(reference_columns) + 1
            for idx, col_name in enumerate(extra_columns, start_col_idx):
                # 写入列名
                temp_ws.cell(row=1, column=idx, value=col_name)
                
                # 获取当前工作表中此列的索引
                curr_col_idx = current_columns.index(col_name) + 1
                
                # 复制数据
                for row_idx in range(1, data_rows + 1):
                    cell_value = ws.cell(row=row_idx + 1 if has_headers else row_idx, 
                                         column=curr_col_idx).value
                    temp_ws.cell(row=row_idx + 1, column=idx, value=cell_value)
                
                print(f"  - 已添加额外列: {col_name}")
        
        # 复制临时工作表的内容回原工作表
        # 首先清空原工作表
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        
        # 复制临时工作表的内容到原工作表
        for row_idx in range(1, temp_ws.max_row + 1):
            for col_idx in range(1, temp_ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx, value=temp_ws.cell(row=row_idx, column=col_idx).value)
        
        # 删除临时工作表
        wb.remove(temp_ws)
        
        print(f"工作表 '{sheet_name}' 的列已重排序")
    
    return wb

def main():
    # 构建输出文件名
    file_name, file_ext = os.path.splitext(input_file)
    output_file = f"{file_name}（已修改）{file_ext}"
    
    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 文件 '{input_file}' 不存在!")
        return
    
    # 打开工作簿
    print(f"正在处理文件: {input_file}")
    wb = openpyxl.load_workbook(input_file)
    
    # 执行列重排序（直接在原工作表中修改）
    wb = reorder_columns_in_place(wb, sheet_settings["sheets_to_process"], sheet_settings["has_headers"])
    
    # 保存修改后的工作簿
    wb.save(output_file)
    print(f"\n列重排序完成! 工作簿已保存为: {output_file}")

if __name__ == "__main__":
    main()