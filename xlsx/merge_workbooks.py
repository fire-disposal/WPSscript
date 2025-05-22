#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
合并多个Excel工作簿
"""

import os
import openpyxl
from datetime import datetime

# 文件读取部分，便于修改需读取文件名
input_files = [
    "workbook1.xlsx",
    "workbook2.xlsx",
    "workbook3.xlsx"
]  # 请修改为实际的文件名列表

# 输出文件名
output_file = f"合并工作簿（{datetime.now().strftime('%Y%m%d_%H%M%S')}）.xlsx"

def merge_workbooks(file_paths):
    """
    合并多个Excel工作簿
    
    Args:
        file_paths: Excel文件路径列表
    
    Returns:
        合并后的Workbook对象
    """
    # 检查文件是否都存在
    missing_files = [f for f in file_paths if not os.path.exists(f)]
    if missing_files:
        print(f"错误: 以下文件不存在: {', '.join(missing_files)}")
        return None
    
    # 创建一个新的工作簿作为合并的目标
    merged_wb = openpyxl.Workbook()
    
    # 删除默认创建的工作表
    default_sheet = merged_wb.active
    merged_wb.remove(default_sheet)
    
    # 记录每个文件的工作表数量
    sheet_counts = {}
    
    # 遍历所有输入文件
    for i, file_path in enumerate(file_paths):
        print(f"正在处理文件 {i+1}/{len(file_paths)}: {file_path}")
        
        # 打开当前工作簿
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 记录工作表数量
        sheet_counts[file_path] = len(wb.sheetnames)
        
        # 获取文件名（不含扩展名）
        file_name = os.path.splitext(os.path.basename(file_path))[0]
        
        # 复制所有工作表
        for sheet_name in wb.sheetnames:
            # 获取源工作表
            source_sheet = wb[sheet_name]
            
            # 创建新的工作表名称（文件名_工作表名）
            new_sheet_name = f"{file_name}_{sheet_name}"
            
            # 如果名称太长，进行截断
            if len(new_sheet_name) > 31:  # Excel工作表名称最大长度为31个字符
                new_sheet_name = new_sheet_name[:28] + "..."
            
            # 确保工作表名称唯一
            counter = 1
            original_name = new_sheet_name
            while new_sheet_name in merged_wb.sheetnames:
                suffix = f"_{counter}"
                # 确保添加后缀后的名称不超过31个字符
                if len(original_name) + len(suffix) > 31:
                    new_sheet_name = original_name[:31-len(suffix)] + suffix
                else:
                    new_sheet_name = original_name + suffix
                counter += 1
            
            # 创建新工作表
            target_sheet = merged_wb.create_sheet(title=new_sheet_name)
            
            # 复制单元格数据
            for row in source_sheet.iter_rows():
                for cell in row:
                    # 获取单元格值
                    value = cell.value
                    
                    # 创建目标单元格
                    target_cell = target_sheet.cell(
                        row=cell.row, 
                        column=cell.column,
                        value=value
                    )
                    
                    # 复制单元格样式
                    if cell.has_style:
                        target_cell.font = cell.font
                        target_cell.border = cell.border
                        target_cell.fill = cell.fill
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection
                        target_cell.alignment = cell.alignment
            
            # 复制列宽
            for col_idx, col in enumerate(source_sheet.columns, 1):
                if source_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width:
                    target_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = \
                        source_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width
            
            # 复制行高
            for row_idx, row in enumerate(source_sheet.rows, 1):
                if source_sheet.row_dimensions[row_idx].height:
                    target_sheet.row_dimensions[row_idx].height = \
                        source_sheet.row_dimensions[row_idx].height
            
            # 复制合并单元格
            for merged_cell_range in source_sheet.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_cell_range))
            
            print(f"  - 已复制工作表: {sheet_name} -> {new_sheet_name}")
        
        print(f"已合并文件: {file_path}")
    
    # 打印合并统计信息
    print("\n合并统计:")
    total_sheets = 0
    for file_path, count in sheet_counts.items():
        print(f"  - {os.path.basename(file_path)}: {count} 个工作表")
        total_sheets += count
    
    print(f"  - 总计: {total_sheets} 个工作表")
    
    return merged_wb

def main():
    # 执行工作簿合并
    merged_wb = merge_workbooks(input_files)
    
    if merged_wb:
        # 保存合并后的工作簿
        merged_wb.save(output_file)
        print(f"合并完成! 工作簿已保存为: {output_file}")
        print(f"共合并了 {len(input_files)} 个工作簿")

if __name__ == "__main__":
    main()