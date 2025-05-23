#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel单元格格式快速处理工具
支持智能格式检测、格式模板应用、条件格式化、格式复制等功能
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.styles.numbers import (
    FORMAT_GENERAL, FORMAT_NUMBER, FORMAT_NUMBER_00, FORMAT_NUMBER_COMMA_SEPARATED1,
    FORMAT_PERCENTAGE, FORMAT_PERCENTAGE_00, FORMAT_DATE_YYYYMMDD2, FORMAT_CURRENCY_USD_SIMPLE
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import re
from datetime import datetime

# 文件读取部分，便于修改需读取文件名
input_file = "example.xlsx"  # 请修改为实际的文件名
output_file = None  # 如果为None，则自动生成输出文件名

# 预定义格式模板
FORMAT_TEMPLATES = {
    "货币": {
        "number_format": "#,##0.00",
        "alignment": {"horizontal": "right", "vertical": "center"},
        "font": {"name": "Arial", "size": 10}
    },
    "百分比": {
        "number_format": "0.00%",
        "alignment": {"horizontal": "center", "vertical": "center"},
        "font": {"name": "Arial", "size": 10}
    },
    "日期": {
        "number_format": "yyyy-mm-dd",
        "alignment": {"horizontal": "center", "vertical": "center"},
        "font": {"name": "Arial", "size": 10}
    },
    "标题": {
        "font": {"name": "Arial", "size": 12, "bold": True},
        "alignment": {"horizontal": "center", "vertical": "center"},
        "fill": {"type": "solid", "color": "DDEBF7"},
        "border": {"style": "thin", "color": "000000"}
    },
    "正文": {
        "font": {"name": "Arial", "size": 10},
        "alignment": {"horizontal": "left", "vertical": "center"},
        "border": {"style": "thin", "color": "000000"}
    },
    "强调": {
        "font": {"name": "Arial", "size": 10, "bold": True, "color": "FF0000"},
        "alignment": {"horizontal": "center", "vertical": "center"},
        "border": {"style": "medium", "color": "000000"}
    }
}

# 格式化操作配置
format_operations = [
    # 操作1：应用预定义模板到特定区域
    {
        "type": "apply_template",
        "sheet": "Sheet1",
        "range": "A1:G1",
        "template": "标题"
    },
    # 操作2：智能格式检测并应用
    {
        "type": "smart_format",
        "sheet": "Sheet1",
        "range": "B2:B20",
        "detect_types": ["number", "date", "percentage", "text"]
    },
    # 操作3：条件格式化
    {
        "type": "conditional_format",
        "sheet": "Sheet1",
        "range": "C2:C20",
        "rules": [
            {
                "type": "cell_value",
                "operator": "greaterThan",
                "formula": "100",
                "font_color": "00FF00",  # 绿色
                "fill_color": "E2EFDA"
            },
            {
                "type": "cell_value",
                "operator": "lessThan",
                "formula": "0",
                "font_color": "FF0000",  # 红色
                "fill_color": "FFCCCC"
            }
        ]
    },
    # 操作4：复制格式
    {
        "type": "copy_format",
        "sheet": "Sheet1",
        "source_range": "A1:G1",
        "target_range": "A10:G10"
    },
    # 操作5：清除格式
    {
        "type": "clear_format",
        "sheet": "Sheet1",
        "range": "D5:D15"
    }
]

# 其他可能的操作示例
"""
# 应用交替行颜色
{
    "type": "alternate_rows",
    "sheet": "Sheet1",
    "range": "A2:G20",
    "even_color": "F2F2F2",
    "odd_color": "FFFFFF"
},
# 设置数据验证
{
    "type": "data_validation",
    "sheet": "Sheet1",
    "range": "E2:E20",
    "validation_type": "list",
    "formula1": "是,否"
}
"""

def create_style_from_template(template_name):
    """
    根据模板名称创建样式对象
    
    Args:
        template_name: 模板名称
    
    Returns:
        样式对象字典
    """
    if template_name not in FORMAT_TEMPLATES:
        print(f"警告: 模板 '{template_name}' 不存在，将使用默认样式")
        return {}
    
    template = FORMAT_TEMPLATES[template_name]
    style = {}
    
    # 创建字体对象
    if "font" in template:
        style["font"] = Font(
            name=template["font"].get("name", "Arial"),
            size=template["font"].get("size", 11),
            bold=template["font"].get("bold", False),
            italic=template["font"].get("italic", False),
            color=template["font"].get("color", "000000")
        )
    
    # 创建填充对象
    if "fill" in template:
        style["fill"] = PatternFill(
            fill_type=template["fill"].get("type", "solid"),
            fgColor=template["fill"].get("color", "FFFFFF")
        )
    
    # 创建对齐对象
    if "alignment" in template:
        style["alignment"] = Alignment(
            horizontal=template["alignment"].get("horizontal", "general"),
            vertical=template["alignment"].get("vertical", "bottom"),
            wrap_text=template["alignment"].get("wrap_text", False)
        )
    
    # 创建边框对象
    if "border" in template:
        border_style = template["border"].get("style", "thin")
        border_color = template["border"].get("color", "000000")
        style["border"] = Border(
            left=Side(style=border_style, color=border_color),
            right=Side(style=border_style, color=border_color),
            top=Side(style=border_style, color=border_color),
            bottom=Side(style=border_style, color=border_color)
        )
    
    # 设置数字格式
    if "number_format" in template:
        style["number_format"] = template["number_format"]
    
    return style

def detect_cell_type(value):
    """
    检测单元格值的类型
    
    Args:
        value: 单元格值
    
    Returns:
        检测到的类型: "number", "percentage", "date", "currency", "text"
    """
    if value is None:
        return "text"
    
    if isinstance(value, (int, float)):
        # 检查是否可能是百分比
        if 0 <= value <= 1:
            return "percentage"
        return "number"
    
    if isinstance(value, datetime):
        return "date"
    
    if isinstance(value, str):
        # 检查是否是货币格式
        if re.match(r'^[$¥€£]', value) or re.search(r'[0-9]+(\.[0-9]{2})?\s*[$¥€£]$', value):
            return "currency"
        
        # 检查是否是百分比格式
        if re.match(r'^[0-9.]+%$', value):
            return "percentage"
        
        # 检查是否是日期格式
        date_patterns = [
            r'\d{4}[-/]\d{1,2}[-/]\d{1,2}',  # yyyy-mm-dd 或 yyyy/mm/dd
            r'\d{1,2}[-/]\d{1,2}[-/]\d{4}',  # dd-mm-yyyy 或 dd/mm/yyyy
            r'\d{1,2}[-/]\d{1,2}[-/]\d{2}'   # dd-mm-yy 或 dd/mm/yy
        ]
        for pattern in date_patterns:
            if re.match(pattern, value):
                return "date"
    
    return "text"

def apply_smart_format(cell, cell_type):
    """
    根据检测到的类型应用智能格式
    
    Args:
        cell: 单元格对象
        cell_type: 检测到的类型
    """
    if cell_type == "number":
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center")
    
    elif cell_type == "percentage":
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    elif cell_type == "date":
        cell.number_format = "yyyy-mm-dd"
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    elif cell_type == "currency":
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center")
    
    else:  # text
        cell.alignment = Alignment(horizontal="left", vertical="center")

def apply_cell_formats(file_path, operations, output_path=None):
    """
    应用单元格格式化操作
    
    Args:
        file_path: 输入Excel文件路径
        operations: 格式化操作列表
        output_path: 输出Excel文件路径，如果为None则自动生成
    
    Returns:
        输出文件路径
    """
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误: 文件 '{file_path}' 不存在!")
        return None
    
    try:
        # 打开工作簿
        print(f"正在打开文件: {file_path}")
        wb = openpyxl.load_workbook(file_path)
        
        # 处理每个操作
        for i, operation in enumerate(operations):
            op_type = operation.get("type")
            sheet_name = operation.get("sheet")
            
            # 检查工作表是否存在
            if sheet_name not in wb.sheetnames:
                print(f"错误: 工作表 '{sheet_name}' 不存在! 跳过操作 #{i+1}")
                continue
            
            ws = wb[sheet_name]
            print(f"正在处理操作 #{i+1}: {op_type} 在工作表 '{sheet_name}'")
            
            # 根据操作类型执行不同的格式化
            if op_type == "apply_template":
                cell_range = ws[operation["range"]]
                template_name = operation["template"]
                style = create_style_from_template(template_name)
                
                for row in cell_range:
                    for cell in row:
                        if "font" in style:
                            cell.font = style["font"]
                        if "fill" in style:
                            cell.fill = style["fill"]
                        if "alignment" in style:
                            cell.alignment = style["alignment"]
                        if "border" in style:
                            cell.border = style["border"]
                        if "number_format" in style:
                            cell.number_format = style["number_format"]
                
                print(f"  - 已应用模板 '{template_name}' 到区域 {operation['range']}")
            
            elif op_type == "smart_format":
                cell_range = ws[operation["range"]]
                detect_types = operation.get("detect_types", ["number", "date", "percentage", "text"])
                
                for row in cell_range:
                    for cell in row:
                        if cell.value is not None:
                            cell_type = detect_cell_type(cell.value)
                            if cell_type in detect_types:
                                apply_smart_format(cell, cell_type)
                
                print(f"  - 已智能格式化区域 {operation['range']}")
            
            elif op_type == "conditional_format":
                cell_range_str = operation["range"]
                rules = operation.get("rules", [])
                
                for rule in rules:
                    rule_type = rule.get("type")
                    
                    if rule_type == "cell_value":
                        operator = rule.get("operator")
                        formula = rule.get("formula")
                        font_color = rule.get("font_color")
                        fill_color = rule.get("fill_color")
                        
                        # 创建条件格式规则
                        cf_rule = CellIsRule(
                            operator=operator,
                            formula=[formula],
                            stopIfTrue=False,
                            font=Font(color=font_color) if font_color else None,
                            fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
                        )
                        
                        # 应用规则
                        ws.conditional_formatting.add(cell_range_str, cf_rule)
                
                print(f"  - 已应用条件格式到区域 {cell_range_str}")
            
            elif op_type == "copy_format":
                source_range = ws[operation["source_range"]]
                target_range = ws[operation["target_range"]]
                
                # 获取源区域和目标区域的行列数
                source_rows = len(source_range)
                source_cols = len(source_range[0]) if source_rows > 0 else 0
                target_rows = len(target_range)
                target_cols = len(target_range[0]) if target_rows > 0 else 0
                
                # 复制格式
                for i in range(min(source_rows, target_rows)):
                    for j in range(min(source_cols, target_cols)):
                        source_cell = source_range[i][j]
                        target_cell = target_range[i][j]
                        
                        target_cell.font = source_cell.font
                        target_cell.border = source_cell.border
                        target_cell.fill = source_cell.fill
                        target_cell.number_format = source_cell.number_format
                        target_cell.alignment = source_cell.alignment
                
                print(f"  - 已从区域 {operation['source_range']} 复制格式到区域 {operation['target_range']}")
            
            elif op_type == "clear_format":
                cell_range = ws[operation["range"]]
                
                for row in cell_range:
                    for cell in row:
                        cell.font = Font()
                        cell.border = Border()
                        cell.fill = PatternFill()
                        cell.number_format = "General"
                        cell.alignment = Alignment()
                
                print(f"  - 已清除区域 {operation['range']} 的格式")
            
            elif op_type == "alternate_rows":
                cell_range = ws[operation["range"]]
                even_color = operation.get("even_color", "F2F2F2")
                odd_color = operation.get("odd_color", "FFFFFF")
                
                for i, row in enumerate(cell_range):
                    fill_color = even_color if i % 2 == 0 else odd_color
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                print(f"  - 已应用交替行颜色到区域 {operation['range']}")
            
            else:
                print(f"  - 未知操作类型: {op_type}")
        
        # 生成输出文件名
        if output_path is None:
            file_name, file_ext = os.path.splitext(file_path)
            output_path = f"{file_name}（已修改）{file_ext}"
        
        # 保存工作簿
        wb.save(output_path)
        print(f"格式化完成! 已保存为: {output_path}")
        
        return output_path
    
    except Exception as e:
        print(f"格式化过程中出错: {str(e)}")
        return None

def main():
    # 生成输出文件名
    global output_file
    if output_file is None:
        file_name, file_ext = os.path.splitext(input_file)
        output_file = f"{file_name}（已修改）{file_ext}"
    
    # 执行格式化操作
    result_file = apply_cell_formats(input_file, format_operations, output_file)
    
    if result_file:
        print(f"Excel单元格格式化成功完成，结果保存到: {result_file}")
    else:
        print("格式化操作失败，请检查输入文件和操作配置。")

if __name__ == "__main__":
    main()
