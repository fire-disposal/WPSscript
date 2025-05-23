#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel工作表行列转置
将Excel工作表中的行和列进行转置（行变列，列变行）
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

# 文件读取部分，便于修改需读取文件名
input_file = "example.xlsx"  # 请修改为实际的文件名
sheet_name = "Sheet1"  # 请修改为实际的工作表名称

# 输出文件设置
output_file = None  # 如果为None，则自动生成输出文件名

def transpose_worksheet(file_path, sheet_name, output_path=None):
    """
    转置Excel工作表的行和列
    
    Args:
        file_path: 输入Excel文件路径
        sheet_name: 要转置的工作表名称
        output_path: 输出Excel文件路径，如果为None则自动生成
    
    Returns:
        输出文件路径
    """
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误: 文件 '{file_path}' 不存在!")
        return None
    
    try:
        # 打开工作簿
        print(f"正在打开文件: {file_path}")
        wb = openpyxl.load_workbook(file_path)
        
        # 检查工作表是否存在
        if sheet_name not in wb.sheetnames:
            print(f"错误: 工作表 '{sheet_name}' 不存在!")
            wb.close()
            return None
        
        # 获取源工作表
        source_sheet = wb[sheet_name]
        
        # 创建一个新的工作簿用于保存转置后的数据
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        output_sheet.title = f"{sheet_name}_转置"
        
        # 获取源工作表的数据范围
        max_row = source_sheet.max_row
        max_col = source_sheet.max_column
        
        print(f"正在转置工作表: {sheet_name} ({max_row}行 x {max_col}列)")
        
        # 执行转置操作
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                # 获取源单元格
                source_cell = source_sheet.cell(row=r, column=c)
                
                # 转置: 行变列，列变行
                output_cell = output_sheet.cell(row=c, column=r)
                
                # 复制值
                output_cell.value = source_cell.value
                
                # 复制样式
                if source_cell.has_style:
                    output_cell.font = source_cell.font
                    output_cell.border = source_cell.border
                    output_cell.fill = source_cell.fill
                    output_cell.number_format = source_cell.number_format
                    output_cell.alignment = source_cell.alignment
        
        # 调整列宽以适应内容
        for col in range(1, max_row + 1):
            output_sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # 生成输出文件名
        if output_path is None:
            file_name, file_ext = os.path.splitext(file_path)
            output_path = f"{file_name}（转置）{file_ext}"
        
        # 保存转置后的工作簿
        output_wb.save(output_path)
        print(f"转置完成! 已保存为: {output_path}")
        
        # 关闭工作簿
        wb.close()
        output_wb.close()
        
        return output_path
    
    except Exception as e:
        print(f"转置过程中出错: {str(e)}")
        return None

def main():
    # 生成输出文件名
    global output_file
    if output_file is None:
        file_name, file_ext = os.path.splitext(input_file)
        output_file = f"{file_name}（转置）{file_ext}"
    
    # 执行转置操作
    result_file = transpose_worksheet(input_file, sheet_name, output_file)
    
    if result_file:
        print(f"工作表 '{sheet_name}' 已成功转置并保存到 '{result_file}'")
        
        # 打印转置前后的对比信息
        try:
            # 打开原始工作簿
            wb_original = openpyxl.load_workbook(input_file)
            original_sheet = wb_original[sheet_name]
            original_rows = original_sheet.max_row
            original_cols = original_sheet.max_column
            
            # 打开转置后的工作簿
            wb_transposed = openpyxl.load_workbook(result_file)
            transposed_sheet = wb_transposed.active
            transposed_rows = transposed_sheet.max_row
            transposed_cols = transposed_sheet.max_column
            
            print("\n转置前后对比:")
            print(f"  - 转置前: {original_rows}行 x {original_cols}列")
            print(f"  - 转置后: {transposed_rows}行 x {transposed_cols}列")
            
            # 关闭工作簿
            wb_original.close()
            wb_transposed.close()
        except Exception as e:
            print(f"获取对比信息时出错: {str(e)}")
    else:
        print("转置操作失败，请检查输入文件和工作表名称。")

if __name__ == "__main__":
    main()